#!/usr/bin/env python3
"""
NewsPortal(DEC) ニュース収集スクリプト

官公庁・関連機関のRSS/Atomフィードを巡回し、脱炭素・水素・再生可能エネルギー
関連の記事のみを抽出して site/data/{articles,feed,categories,topics,ranks}.json
を更新する。GitHub Actions から3時間おきに実行される想定。

あわせて、キー登録不要で取得できる外部公開データを使って以下も実数値で更新する
(取得に失敗した場合は前回値を保持し、エラーにはしない):
  - 環境省サイトの「ゼロカーボンシティ」表明自治体数 → kpis.json / dashboard.json
  - 環境省の表明自治体「取組一覧」PDFから集計した都道府県別の表明件数
    → dashboard.json の zeroCarbonByPrefecture
    (いずれも日々の変動が少ないため週1回のみ収集し、収集日をasOfに明示する)
  - JEPX(日本卸電力取引所)スポット市場CSVのシステムプライス平均 → rail-data.json

events.json / glossary.json / shortcuts.json、および dashboard.json の地域別
(8地域)ゼロカーボン内訳・電源構成・CO2排出量トレンド等は、統一的に取得できる
公開API/ファイルが確認できなかったため対象外（静的データのまま）としている。
"""
import csv
import hashlib
import io
import json
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone, timedelta
from pathlib import Path
from xml.etree import ElementTree

BASE_DIR = Path(__file__).resolve().parent
SITE_DATA_DIR = BASE_DIR.parent / "site" / "data"
SOURCES_FILE = BASE_DIR / "sources.json"

JST = timezone(timedelta(hours=9))
WEEKDAY_JA = ["月", "火", "水", "木", "金", "土", "日"]

MAX_ARTICLES = 1000
MAX_FEED_ITEMS = 30
FETCH_TIMEOUT = 15
USER_AGENT = "Mozilla/5.0 (compatible; NewsPortalDEC-Collector/1.0)"

# 脱炭素・水素・再エネ関連かどうかを判定するキーワード
KEYWORDS = [
    "脱炭素", "カーボン", "CO2", "CO₂", "温室効果ガス", "GHG", "GX",
    "再生可能エネルギー", "再エネ", "太陽光", "風力", "洋上風力", "地熱",
    "水素", "燃料電池", "FCV", "アンモニア", "蓄電池", "蓄電", "EV",
    "ゼロカーボン", "カーボンニュートラル", "省エネ", "排出量取引",
    "Jクレジット", "J-クレジット", "SAF", "e-fuel", "合成燃料",
]
KEYWORD_RE = re.compile("|".join(re.escape(k) for k in KEYWORDS))
KEYWORDS_SET = set(KEYWORDS)

# カテゴリ自動分類（一致したら上書き、優先順位は上から）
CATEGORY_RULES = [
    (re.compile("水素|燃料電池|FCV|アンモニア"), "技術"),
    # 「地域」は「先行地域」(脱炭素先行地域)を誤って自治体カテゴリに引き込んでしまうため除外する
    (re.compile("自治体|地方公共団体|ゼロカーボンシティ|(?<!先行)地域|港湾|都市"), "自治体"),
    (re.compile("海外|米国|アメリカ|欧州|ヨーロッパ|英国|イギリス|ドイツ|フランス|中国|インド|韓国|台湾|アジア|中東|国連|世界|COP\d+"), "国際"),
    (re.compile("住宅|家庭|くらし|生活者|家電|節電|電気料金|ZEH"), "暮らし"),
]

# 「国」カテゴリは政府系の報道発表のみに限定する。引き継ぎデータ(articles.json)に
# 過去の分類ルールなどで「国」として残っている他発行元の記事は再分類の対象とする。
APPROVED_NATIONAL_SOURCES = {"環境省", "経済産業省", "資源エネルギー庁"}

XML_NS = {
    "rdf": "http://www.w3.org/1999/02/22-rdf-syntax-ns#",
    "rss1": "http://purl.org/rss/1.0/",
    "atom": "http://www.w3.org/2005/Atom",
    "dc": "http://purl.org/dc/elements/1.1/",
}

GOOGLE_NEWS_RSS = "https://news.google.com/rss/search"
# Googleニュース検索RSSのタイトルは「本文 - 配信元」の形式になっている
TITLE_SOURCE_SUFFIX_RE = re.compile(r"^(?P<title>.+?)\s+-\s+(?P<source>[^-]+)$")


def build_search_url(query: str) -> str:
    """検索キーワードから Google ニュース検索RSSのURLを組み立てる"""
    params = urllib.parse.urlencode({"q": query, "hl": "ja", "gl": "JP", "ceid": "JP:ja"})
    return f"{GOOGLE_NEWS_RSS}?{params}"


def split_title_source(raw_title: str, fallback: str):
    m = TITLE_SOURCE_SUFFIX_RE.match(raw_title)
    if m:
        return m.group("title").strip(), m.group("source").strip()
    return raw_title.strip(), fallback


def fetch(url: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=FETCH_TIMEOUT) as res:
        return res.read()


def strip_html(text: str) -> str:
    text = re.sub(r"<[^>]+>", "", text or "")
    return re.sub(r"\s+", " ", text).strip()


def parse_feed(xml_bytes: bytes):
    """RSS2.0 / RDF(RSS1.0) / Atom を雑にまとめてパースし (title, link, desc, pubdate) を返す"""
    items = []
    try:
        root = ElementTree.fromstring(xml_bytes)
    except ElementTree.ParseError:
        return items

    tag = root.tag.lower()

    if tag.endswith("rdf"):
        for item in root.findall("rss1:item", XML_NS):
            title = item.findtext("rss1:title", default="", namespaces=XML_NS)
            link = item.findtext("rss1:link", default="", namespaces=XML_NS)
            desc = item.findtext("rss1:description", default="", namespaces=XML_NS)
            date = item.findtext("dc:date", default="", namespaces=XML_NS)
            items.append((title, link, desc, date))
    elif tag.endswith("feed"):  # Atom
        for entry in root.findall("atom:entry", XML_NS):
            title = entry.findtext("atom:title", default="", namespaces=XML_NS)
            link_el = entry.find("atom:link", XML_NS)
            link = link_el.get("href") if link_el is not None else ""
            desc = entry.findtext("atom:summary", default="", namespaces=XML_NS)
            date = entry.findtext("atom:updated", default="", namespaces=XML_NS)
            items.append((title, link, desc, date))
    else:  # RSS 2.0
        for item in root.iter("item"):
            title = item.findtext("title", default="")
            link = item.findtext("link", default="")
            desc = item.findtext("description", default="")
            date = item.findtext("pubDate", default="")
            items.append((title, link, desc, date))

    return items


PRESS_LIST_BLOCK_RE = re.compile(
    r'<span class="p-press-release-list__heading">([^<]+)</span>(.*?)</details>',
    re.S,
)
PRESS_LIST_ITEM_RE = re.compile(r'<a href="([^"]+)" class="c-news-link__link">([^<]+)</a>')


def parse_press_html_list(html_bytes: bytes, base_url: str):
    """RSSを提供していない環境省 報道発表一覧ページ(HTML)から (title, link, desc, date_raw) を抜き出す"""
    html_text = html_bytes.decode("utf-8", errors="ignore")
    items = []
    for date_raw, body in PRESS_LIST_BLOCK_RE.findall(html_text):
        for href, title in PRESS_LIST_ITEM_RE.findall(body):
            link = urllib.parse.urljoin(base_url, href)
            items.append((title, link, "", date_raw))
    return items


# 「脱炭素先行地域づくり支援サイト」の先行地域ページ自体には新着記事一覧がなく、
# 「脱炭素先行地域評価委員会」の開催状況一覧だけが日付付きで定期的に更新されるため、
# これを先行地域カテゴリの最新ニュース相当として抜き出す
PRECEDING_REGION_COMMITTEE_RE = re.compile(
    r'<a href="(?P<href>[^"]+)"[^>]*>\s*第(?P<round>[0-9０-９]+)回\s*'
    r'脱炭素先行地域評価委員会\s*</a>\s*<br>\s*'
    r'（日時：令和(?P<era_year>[0-9０-９]+)年(?P<month>[0-9０-９]+)月(?P<day>[0-9０-９]+)日'
    # 通常は「／場所：○○」だが、書面開催の回は「／書面開催」のみで「場所：」が付かない
    r'／(?:場所：)?(?P<place>[^）]*)）'
)
ZENKAKU_DIGITS = str.maketrans("０１２３４５６７８９", "0123456789")


def parse_preceding_region_committee(html_bytes: bytes, base_url: str):
    """脱炭素先行地域づくり支援サイトの「開催状況」一覧から (title, link, desc, date_raw) を
    抜き出す。年月日は元号(令和)表記かつ全角/半角数字が混在しているため、半角化したうえで
    西暦に変換する"""
    html_text = html_bytes.decode("utf-8", errors="ignore")
    items = []
    for m in PRECEDING_REGION_COMMITTEE_RE.finditer(html_text):
        round_no = m.group("round").translate(ZENKAKU_DIGITS)
        era_year = int(m.group("era_year").translate(ZENKAKU_DIGITS))
        month = int(m.group("month").translate(ZENKAKU_DIGITS))
        day = int(m.group("day").translate(ZENKAKU_DIGITS))
        year = era_year + 2018  # 令和1年 = 2019年
        title = f"脱炭素先行地域評価委員会（第{round_no}回）を開催"
        link = urllib.parse.urljoin(base_url, m.group("href"))
        desc = f"場所：{m.group('place').strip()}"
        items.append((title, link, desc, f"{year}-{month:02d}-{day:02d}"))
    return items


# 「脱炭素先行地域中間評価」セクション(id=chukan)。各自治体ごとの進捗報告書まで含む
# 表は対象外とし、令和X年度ごとの中間評価結果(総評/総評について)のみを抜き出す
PRECEDING_REGION_CHUKAN_ITEM_RE = re.compile(
    r'<td>(?P<title>令和[0-9０-９]+年度脱炭素先行地域中間評価[^<]*)</td>\s*'
    r'<td[^>]*>\s*<a href="(?P<href>[^"]+\.pdf)"'
)
# 一覧ページ自体に公表日の記載がないため、各PDF本文1ページ目に印字されている
# 公表日(令和表記)を直接読み取る。PDFからのテキスト抽出では「令和 ７ 年 ２ 月 27 日」
# のように桁や単位の前後に余分な空白が入ることがあるため、それらすべてを許容して読み取る
_DIGITS = r"[0-9０-９](?:\s*[0-9０-９])*"
PDF_PUBLISH_DATE_RE = re.compile(
    rf"令和\s*(?P<era_year>{_DIGITS})\s*年\s*(?P<month>{_DIGITS})\s*月\s*(?P<day>{_DIGITS})\s*日"
)


def extract_pdf_publish_date(pdf_bytes: bytes, context: str = ""):
    """PDF1ページ目のテキストから令和表記の公表日を西暦(YYYY-MM-DD)で返す(抽出失敗時はNone)"""
    if pypdf is None:
        return None
    try:
        reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
        text = reader.pages[0].extract_text() or ""
    except Exception as exc:
        print(f"[diag] PDF解析失敗 ({context}): {exc}", file=sys.stderr)
        return None
    m = PDF_PUBLISH_DATE_RE.search(text)
    if not m:
        print(f"[diag] 公表日の正規表現が不一致 ({context}): 抽出テキスト先頭200文字={text[:200]!r}", file=sys.stderr)
        return None

    def digits(group: str) -> int:
        return int(re.sub(r"\s+", "", group).translate(ZENKAKU_DIGITS))

    era_year, month, day = digits(m.group("era_year")), digits(m.group("month")), digits(m.group("day"))
    return f"{era_year + 2018}-{month:02d}-{day:02d}"


def parse_preceding_region_chukan(html_bytes: bytes, base_url: str):
    """脱炭素先行地域づくり支援サイトの「中間評価」セクション(id=chukan)から資料一覧を
    (title, link, desc, date_raw) として抜き出す。一覧ページに日付の記載がないため、
    各PDFを取得して公表日を読み取る(PDF取得・解析に失敗した場合は date_raw を
    Noneとし、収集時刻が日付として採用される)"""
    html_text = html_bytes.decode("utf-8", errors="ignore")
    section_start = html_text.find('id="chukan"')
    if section_start == -1:
        return []
    section_end = html_text.find('id="progress"', section_start)
    section = html_text[section_start:section_end if section_end != -1 else None]

    items = []
    for m in PRECEDING_REGION_CHUKAN_ITEM_RE.finditer(section):
        title = m.group("title")
        link = urllib.parse.urljoin(base_url, m.group("href"))
        date_raw = None
        try:
            date_raw = extract_pdf_publish_date(fetch(link), context=link)
        except (urllib.error.URLError, TimeoutError, ValueError) as exc:
            print(f"[skip] 先行地域中間評価PDFの日付取得失敗 ({link}): {exc}", file=sys.stderr)
        items.append((title, link, "", date_raw))
    return items


HTML_LIST_PARSERS = {
    "press_release_list": parse_press_html_list,
    "preceding_region_committee": parse_preceding_region_committee,
    "preceding_region_chukan": parse_preceding_region_chukan,
}


def parse_date(raw: str):
    if not raw:
        return None
    fmts = [
        "%a, %d %b %Y %H:%M:%S %z",
        "%a, %d %b %Y %H:%M:%S %Z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y年%m月%d日発表",
    ]
    for fmt in fmts:
        try:
            dt = datetime.strptime(raw.strip(), fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=JST)
            return dt.astimezone(JST)
        except ValueError:
            continue
    return None


def format_time(dt: datetime) -> str:
    wd = WEEKDAY_JA[dt.weekday()]
    return f"{dt.month}/{dt.day}({wd}) {dt.hour}:{dt.minute:02d}"


def make_id(key: str) -> str:
    return "auto-" + hashlib.sha1(key.encode("utf-8")).hexdigest()[:12]


# 同じ記事が複数のRSS/検索クエリ経由で別リンク・別配信元として重複収集されるのを防ぐための
# タイトル正規化キー。末尾の「(配信元テレビ局)」「(掲載日)」「(3ページ目)」のような
# Yahoo!ニュース等が付与する注記や、全角/半角スペースの差異を吸収して同一記事と判定する
TITLE_DEDUP_SUFFIX_RE = re.compile(r"\s*[\(（][^()（）]{1,30}[\)）]\s*$")


def normalize_title_for_dedup(title: str) -> str:
    return re.sub(r"\s+", "", TITLE_DEDUP_SUFFIX_RE.sub("", title))


def dedupe_by_title(articles: dict) -> dict:
    """正規化タイトルが一致する記事(別IDで重複登録されている過去収集分を含む)を1件に
    統合する。注記がない分タイトルが短くなる傾向を利用し、最も短いタイトルの記事を残す
    (同じ長さなら辞書の並び順で先に見つかった方、すなわちより古い記事を残す)"""
    best = {}
    for a in articles.values():
        key = normalize_title_for_dedup(a["title"])
        current = best.get(key)
        if current is None or len(a["title"]) < len(current["title"]):
            best[key] = a
    return {a["id"]: a for a in best.values()}


def classify_category(text: str, fallback: str) -> str:
    for pattern, category in CATEGORY_RULES:
        if pattern.search(text):
            return category
    return fallback


def reclassify_carried_over_national(carried_over: dict) -> None:
    """引き継ぎ記事のうち、政府系3省庁以外の発行元で「国」のまま残っているものを
    現行のCATEGORY_RULESで再判定し、該当なしなら「その他」に変更する"""
    for article in carried_over.values():
        if article.get("category") != "国" or article.get("source") in APPROVED_NATIONAL_SOURCES:
            continue
        haystack = article.get("title", "") + " " + article.get("summary", "")
        article["category"] = classify_category(haystack, "その他")


IMAGE_FETCH_TIMEOUT = 8
MAX_IMAGE_FETCH_BYTES = 200_000
META_IMAGE_PATTERNS = [
    re.compile(r'<meta[^>]+property=["\']og:image["\'][^>]+content=["\']([^"\']+)["\']', re.I),
    re.compile(r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+property=["\']og:image["\']', re.I),
    re.compile(r'<meta[^>]+name=["\']twitter:image["\'][^>]+content=["\']([^"\']+)["\']', re.I),
    re.compile(r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+name=["\']twitter:image["\']', re.I),
]


def extract_og_image(html_text: str):
    """HTML文字列からog:image(なければtwitter:image)のURLを抜き出す"""
    for pattern in META_IMAGE_PATTERNS:
        m = pattern.search(html_text)
        if m:
            return m.group(1).strip()
    return None


def fetch_article_image(url: str):
    """記事の元ページを取得し、OGP画像(og:image / twitter:image)のURLを返す(失敗時はNone)"""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        with urllib.request.urlopen(req, timeout=IMAGE_FETCH_TIMEOUT) as res:
            content_type = res.headers.get("Content-Type", "")
            if "html" not in content_type.lower():
                return None
            raw = res.read(MAX_IMAGE_FETCH_BYTES)
            final_url = res.geturl()
    except (urllib.error.URLError, TimeoutError, ValueError, OSError) as exc:
        print(f"[skip] 画像取得失敗 ({url}): {exc}", file=sys.stderr)
        return None

    image_url = extract_og_image(raw.decode("utf-8", errors="ignore"))
    if not image_url:
        return None
    return urllib.parse.urljoin(final_url, image_url)


def collect_one_source(source: dict, now: datetime):
    results = []
    source_type = source.get("type", "feed")
    url = build_search_url(source["query"]) if source_type == "search" else source["url"]

    try:
        raw = fetch(url)
    except (urllib.error.URLError, TimeoutError, ValueError) as exc:
        print(f"[skip] {source['name']}: 取得失敗 ({exc})", file=sys.stderr)
        return results

    if source_type == "html_list":
        parser = HTML_LIST_PARSERS[source.get("parser", "press_release_list")]
        items = parser(raw, url)
    else:
        items = parse_feed(raw)

    for raw_title, link, desc, date_raw in items:
        raw_title = strip_html(raw_title)
        desc = strip_html(desc)
        if not raw_title or not link:
            continue

        if source_type == "search":
            # 検索RSSはGoogleニュースの仲介リンクなので「本文 - 配信元」を分離する
            title, source_label = split_title_source(raw_title, source.get("default_source_label", source["name"]))
        else:
            title, source_label = raw_title, source["default_source_label"]

        haystack = title + " " + desc
        if not KEYWORD_RE.search(haystack):
            continue  # 脱炭素・水素・再エネに無関係な記事は除外

        dt = parse_date(date_raw) or now
        matched_tags = list(dict.fromkeys(KEYWORD_RE.findall(haystack)))[:4]

        results.append({
            "id": make_id(normalize_title_for_dedup(title)),
            "category": classify_category(haystack, source["category"]),
            "title": title,
            "source": source_label,
            "source_url": link,
            "time": format_time(dt),
            "_sort_key": dt.isoformat(),
            "comments": 0,
            "tags": matched_tags,
            "summary": (desc[:140] + "…") if len(desc) > 140 else (desc or title),
            "body": [desc] if desc else [title],
            "related": [],
            "collected": True,
        })

    return results


def build_topics(articles: dict, ordered_ids: list, categories: list, new_ids: set):
    """直近の収集記事から、主要トピック(トップ記事+見出しリスト)を組み立てる"""
    tabs = ["総合"] + [c["label"] for c in categories]

    if not ordered_ids:
        return None

    lead_id = ordered_ids[0]
    lead_a = articles[lead_id]
    lead = {
        "id": lead_id,
        "title": lead_a["title"],
        "summary": lead_a["summary"],
        "category": lead_a["category"],
        "source": lead_a["source"],
        "time": lead_a["time"],
        "comments": 0,
        "source_url": lead_a["source_url"],
        "image": lead_a.get("image"),
    }

    headlines = []
    for rank, aid in enumerate(ordered_ids[1:9], start=1):
        a = articles[aid]
        headlines.append({
            "id": aid,
            "rank": rank,
            "title": a["title"],
            "category": a["category"],
            "source": a["source"],
            "time": a["time"],
            "tag": "NEW" if aid in new_ids else "",
            "pr": False,
            "source_url": a["source_url"],
        })

    return {"tabs": tabs, "lead": lead, "headlines": headlines}


# 「話題のキーワード」ランキング専用の特化辞書。記事の関連性判定(KEYWORDS)用の
# 広いカテゴリ語とは別に、環境政策・施策に携わる人が見て分かる固有名詞的な
# 法律・制度・技術名のみを集めている。ランキングの語彙はこれに加えて、下の
# パターンベース抽出で一定の頻度・出典数に達した語を自動的に取り込んでいくため、
# 手動でリストを増やし続けなくても新しい制度・技術名が拾えるようになっている。
RANK_BASE_KEYWORDS = [
    "GX推進法", "GX推進戦略", "地球温暖化対策計画", "エネルギー基本計画",
    "水素基本戦略", "カーボンプライシング", "炭素税", "排出量取引制度",
    "GX-ETS", "Jクレジット", "CBAM", "FIT制度", "FIP制度",
    "RE100", "SBTi", "TCFD", "CCS", "CCUS", "DAC", "SAF",
    "アンモニア混焼", "ペロブスカイト太陽電池", "バイオマス", "グリーン水素",
]

RANK_CANDIDATE_SUFFIXES = [
    "法", "計画", "制度", "戦略", "構想", "宣言", "ロードマップ", "ガイドライン", "イニシアチブ",
]
# 英字略称(RE100・GX-ETSのような表記)。会社名・媒体名(ENEOS・NEWS等)との
# 区別がつかない素の英字のみの語を誤って拾わないよう、数字またはハイフンを
# 含むものだけを候補として扱う(末尾の小文字1字(SBTiのような表記)は許容)
RANK_ACRONYM_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:[A-Z][A-Z0-9]{1,7}-[A-Z0-9]{1,8}|[A-Z][A-Z0-9]{2,7})[a-z]?(?![A-Za-z0-9])"
)
# 「法」は「法人」の一部に偶然マッチしてしまうため、直後が「人」の場合は除外する
RANK_KANJI_SUFFIX_RE = re.compile(
    r"[一-龠ァ-ヴー]{2,14}?(?:法(?!人)|計画|制度|戦略|構想|宣言|ロードマップ|ガイドライン|イニシアチブ)"
)
RANK_PROMOTE_MIN_COUNT = 3
RANK_PROMOTE_MIN_SOURCES = 2


def extract_rank_candidates(text: str) -> set:
    """記事本文から、ランキング語彙の自動候補(数字/ハイフンを含む英字略称・
    法令/制度名らしい漢字複合語)を抜き出す。既存のKEYWORDS(関連性判定用の
    広い語)に含まれるものは、専門ランキング向けの新規候補としては扱わない"""
    found = {
        m for m in RANK_ACRONYM_RE.findall(text)
        if any(c.isdigit() for c in m) or "-" in m
    }
    found.update(RANK_KANJI_SUFFIX_RE.findall(text))
    found -= KEYWORDS_SET
    return found


def build_rank_vocab(articles: dict):
    """直近の記事群(articles.jsonの引き継ぎ分含む全件)から自動候補語を集計し、
    一定の頻度・出典数に達したものだけをベース辞書に加えて返す。
    記事の引き継ぎ自体がローリングウィンドウになっているため、話題性が薄れた語は
    追って自然に閾値を下回り、明示的なプルーニング処理なしで自動的に外れていく"""
    candidate_counts = {}
    candidate_sources = {}
    for a in articles.values():
        haystack = a.get("title", "") + " " + a.get("summary", "")
        for term in extract_rank_candidates(haystack):
            candidate_counts[term] = candidate_counts.get(term, 0) + 1
            candidate_sources.setdefault(term, set()).add(a.get("source", ""))

    promoted = [
        term for term, count in candidate_counts.items()
        if count >= RANK_PROMOTE_MIN_COUNT
        and len(candidate_sources[term]) >= RANK_PROMOTE_MIN_SOURCES
    ]

    return list(dict.fromkeys(RANK_BASE_KEYWORDS + sorted(promoted)))


def build_ranks(articles: dict, previous_ranks: list, top_n: int = 8):
    """環境政策・施策の専門語(ベース辞書＋自動採用された候補語)が記事本文(タイトル+
    概要)に出現する頻度からトレンドキーワードを集計する"""
    vocab = build_rank_vocab(articles)
    if not vocab:
        return None
    vocab_re = re.compile("|".join(re.escape(v) for v in sorted(vocab, key=len, reverse=True)))

    counts = {}
    for a in articles.values():
        haystack = a.get("title", "") + " " + a.get("summary", "")
        for term in set(vocab_re.findall(haystack)):
            counts[term] = counts.get(term, 0) + 1

    if not counts:
        return None

    ranked = sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))[:top_n]
    prev_index = {r["keyword"]: i for i, r in enumerate(previous_ranks)}

    result = []
    for i, (keyword, _count) in enumerate(ranked):
        if keyword not in prev_index:
            trend = "NEW"
        elif prev_index[keyword] - i >= 3:
            trend = "急上昇"
        else:
            trend = ""
        result.append({"keyword": keyword, "trend": trend})
    return result


EVENT_KEYWORDS = [
    "セミナー", "シンポジウム", "フォーラム", "ウェビナー", "説明会",
    "勉強会", "講演会", "展示会", "総会", "カンファレンス", "EXPO", "見本市",
]
EVENT_KEYWORD_RE = re.compile("|".join(EVENT_KEYWORDS))
# 「7月29日に開催」「7月8日(水)開催」のように、日付の直後(住所・会場等の短い修飾を挟むことはある)に
# 「開催」が続く箇所のみを開催日とみなす。記事タイトルに混在する掲載日(「6月27日掲載」等)との
# 誤判定を避けるため。
EVENT_DATE_RE = re.compile(r"(\d{1,2})月(\d{1,2})日.{0,15}?開催")


def extract_event_date(text: str, base: datetime):
    m = EVENT_DATE_RE.search(text)
    if not m:
        return None
    month, day = int(m.group(1)), int(m.group(2))
    try:
        candidate = base.replace(month=month, day=day, hour=0, minute=0, second=0, microsecond=0)
    except ValueError:
        return None
    if candidate.date() < base.date():
        try:
            candidate = candidate.replace(year=candidate.year + 1)
        except ValueError:
            return None
    return candidate


def build_events(articles: dict, previous_events: list, now: datetime, top_n: int = 8):
    """収集記事のうち、セミナー/イベント関連かつ開催日が明記されているものから一覧を組み立てる。
    開催日を過ぎたものは(引き継ぎ分も含めて)除外する"""
    found = {}
    for a in articles.values():
        haystack = a["title"] + " " + " ".join(a.get("body", []))
        if not EVENT_KEYWORD_RE.search(haystack):
            continue
        event_dt = extract_event_date(a["title"], now) or extract_event_date(haystack, now)
        if event_dt is None:
            continue
        key = (event_dt.month, event_dt.day, a["title"])
        found[key] = {
            "month": f"{event_dt.month:02d}",
            "day": f"{event_dt.day:02d}",
            "title": a["title"],
            "place": "オンライン" if "オンライン" in haystack else "",
            "source": a["source"],
            "source_url": a["source_url"],
            "_date": event_dt.isoformat(),
        }

    for e in previous_events:
        raw_date = e.get("_date")
        if not raw_date:
            continue
        try:
            d = datetime.fromisoformat(raw_date)
        except ValueError:
            continue
        if d < now:
            continue
        key = (int(e["month"]), int(e["day"]), e["title"])
        found.setdefault(key, e)

    ordered = sorted(found.values(), key=lambda e: e["_date"])[:top_n]
    for e in ordered:
        e.pop("_date", None)
    return ordered


ZERO_CARBON_URL = "https://www.env.go.jp/policy/zerocarbon.html"
# ゼロカーボンシティ宣言自治体数は日々の変動が少ないため、収集は週1回に間引く。
ZERO_CARBON_COLLECT_INTERVAL_DAYS = 7
ZERO_CARBON_ASOF_RE = re.compile(r"(\d+)年(\d+)月(\d+)日")
# ページ本文には総数の記載がなく、リンク先の「一覧図」PDFに埋め込まれていることが
# 実際のページ調査で判明した。まずページ本文を試し、駄目なら一覧図PDFのテキストを試す。
# 最終的にもありえない値は採用しない(ZERO_CARBON_PLAUSIBLE_RANGE)。
ZERO_CARBON_STRICT_RE = re.compile(r"表明した地方公共団体は.{0,60}?(\d[\d,]{2,})\s*(?:自治体|団体)")
ZERO_CARBON_PDF_COUNT_RE = re.compile(r"表明自治体数\D{0,10}?(\d[\d,]{2,})")
# 一覧図PDFの実際の文言「〇〇を始めとする1215自治体(...)が...を表明。」に合わせて、
# 「を始めとする」直後の数値だけを拾う。単純な「数値+団体」のみだと無関係な箇所
# (例:本文中の年号と「団体」が偶然隣接する箇所)に誤マッチするため。
ZERO_CARBON_LOOSE_RE = re.compile(r"を始めとする(\d[\d,]{2,})\s*(?:自治体|団体)")
ZERO_CARBON_PDF_LINK_RE = re.compile(r'<a\b[^>]*href="([^"]+)"[^>]*>([^<]{0,80})')
ZERO_CARBON_PLAUSIBLE_RANGE = (300, 1800)

JEPX_DOWNLOAD_URL = "https://www.jepx.jp/_download.php"
JEPX_SPOT_PAGE_URL = "https://www.jepx.jp/electricpower/market-data/spot/"
# 補助金・公募情報の収集対象ページ
SUBSIDY_SOURCES_CFG = [
    {
        "name": "環境省 公募情報",
        "url": "https://www.env.go.jp/guide/kobo.html",
        "source_label": "環境省",
        "base_url": "https://www.env.go.jp",
        # このページは「公募中の案件名」テーブルの本文リンクのみが対象案件で、
        # ページ全体には共通ナビ(ホーム/申請手続等)のリンクが大量に含まれるため、
        # <table>タグ内のリンクだけに絞り込む(href_filterでは共通ナビを除外できない)
        "table_scope": True,
    },
    {
        "name": "NEDO 公募情報",
        "url": "https://www.nedo.go.jp/koubo/index.html",
        "source_label": "NEDO",
        "base_url": "https://www.nedo.go.jp",
        "href_filter": "/koubo/",
    },
    {
        "name": "経済産業省 公募情報",
        "url": "https://www.meti.go.jp/information/publicoffer/kobo/index.html",
        "source_label": "経済産業省",
        "base_url": "https://www.meti.go.jp",
        "href_filter": "/information/publicoffer/",
    },
]

_SUBSIDY_A_RE = re.compile(r'<a\b[^>]*href="([^"]*)"[^>]*>\s*([^<]{4,120})\s*</a>', re.S)
_SUBSIDY_NOISE_RE = re.compile(
    r'^\s*(?:ページの先頭へ|前(?:のページ)?|次(?:のページ)?|一覧に戻る|ホーム|トップ(?:ページ)?|'
    r'もっと見る|English|サイトマップ|\d{1,4}(?:件)?|▲|▼|›|»)\s*$'
)

try:
    import pypdf
except Exception:
    pypdf = None

try:
    import openpyxl
except Exception:
    openpyxl = None


def extract_zero_carbon_total_from_text(text: str):
    m = (
        ZERO_CARBON_STRICT_RE.search(text)
        or ZERO_CARBON_PDF_COUNT_RE.search(text)
        or ZERO_CARBON_LOOSE_RE.search(text)
    )
    if not m:
        return None
    total = int(m.group(1).replace(",", ""))
    lo, hi = ZERO_CARBON_PLAUSIBLE_RANGE
    if not (lo <= total <= hi):
        print(f"[skip] ゼロカーボン自治体数: 抽出値 {total} が妥当な範囲外", file=sys.stderr)
        return None
    return total


def find_zero_carbon_pdf_url(html_text: str):
    """ページ内のリンクから「一覧図」PDFのURLを探す"""
    for href, link_text in ZERO_CARBON_PDF_LINK_RE.findall(html_text):
        if "一覧図" in link_text:
            return urllib.parse.urljoin(ZERO_CARBON_URL, href)
    return None


def find_zero_carbon_list_pdf_url(html_text: str):
    """ページ内のリンクから表明自治体の都道府県別「取組一覧」PDFのURLを探す"""
    for href, link_text in ZERO_CARBON_PDF_LINK_RE.findall(html_text):
        if "取組一覧" in link_text:
            return urllib.parse.urljoin(ZERO_CARBON_URL, href)
    return None


ZERO_CARBON_PREFECTURES = [
    "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
    "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県",
    "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県", "岐阜県",
    "静岡県", "愛知県", "三重県",
    "滋賀県", "京都府", "大阪府", "兵庫県", "奈良県", "和歌山県",
    "鳥取県", "島根県", "岡山県", "広島県", "山口県",
    "徳島県", "香川県", "愛媛県", "高知県",
    "福岡県", "佐賀県", "長崎県", "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県",
]

# 各都道府県の市区町村数(東京都は特別区を含む)。市町村合併以外では変動しない
# 静的な行政区分データのため、ハードコードしている(合計1,741市区町村と整合)。
ZERO_CARBON_MUNICIPALITY_TOTALS = {
    "北海道": 179, "青森県": 40, "岩手県": 33, "宮城県": 35, "秋田県": 25, "山形県": 35, "福島県": 59,
    "茨城県": 44, "栃木県": 25, "群馬県": 35, "埼玉県": 63, "千葉県": 54, "東京都": 62, "神奈川県": 33,
    "新潟県": 30, "富山県": 15, "石川県": 19, "福井県": 17, "山梨県": 27, "長野県": 77, "岐阜県": 42,
    "静岡県": 35, "愛知県": 54, "三重県": 29,
    "滋賀県": 19, "京都府": 26, "大阪府": 43, "兵庫県": 41, "奈良県": 39, "和歌山県": 30,
    "鳥取県": 19, "島根県": 19, "岡山県": 27, "広島県": 23, "山口県": 19,
    "徳島県": 24, "香川県": 17, "愛媛県": 20, "高知県": 34,
    "福岡県": 60, "佐賀県": 20, "長崎県": 21, "熊本県": 45, "大分県": 18, "宮崎県": 26, "鹿児島県": 43, "沖縄県": 41,
}

# PDF内の各エントリは「\n<連番> <都道府県名>\n<ふりがな>\n」で始まり、続く行は
# 市区町村名(市区町村による表明)か表明日の日付(都道府県自身による表明)のいずれか
# になる。ZERO_CARBON_MUNICIPALITY_TOTALSは市区町村数のみの集計のため、都道府県
# 自身の表明(続く行が数字=日付で始まるケース)は市区町村の表明として数えない。
ZERO_CARBON_ENTRY_RE = re.compile(
    r"\n\d+\s(" + "|".join(ZERO_CARBON_PREFECTURES) + r")\n[ぁ-んー]+\n(?!\d)"
)


def count_zero_carbon_by_prefecture(pdf_text: str):
    """取組一覧PDFの全文テキストから都道府県ごとの表明件数を集計する"""
    counts = {pref: 0 for pref in ZERO_CARBON_PREFECTURES}
    for m in ZERO_CARBON_ENTRY_RE.finditer(pdf_text):
        counts[m.group(1)] += 1
    return counts


def fetch_zero_carbon_by_prefecture():
    """都道府県別「取組一覧」PDFから表明自治体数を集計する(失敗時はNone)"""
    try:
        raw = fetch(ZERO_CARBON_URL)
    except (urllib.error.URLError, TimeoutError, ValueError) as exc:
        print(f"[skip] ゼロカーボン都道府県別: ページ取得失敗 ({exc})", file=sys.stderr)
        return None

    html_raw_text = raw.decode("utf-8", errors="ignore")
    pdf_url = find_zero_carbon_list_pdf_url(html_raw_text)
    if not pdf_url:
        print("[skip] ゼロカーボン都道府県別: 取組一覧PDFへのリンクが見つからず", file=sys.stderr)
        return None

    pdf_text = fetch_zero_carbon_pdf_text(pdf_url)
    if not pdf_text:
        print("[skip] ゼロカーボン都道府県別: PDFテキスト抽出失敗", file=sys.stderr)
        return None

    counts = count_zero_carbon_by_prefecture(pdf_text)
    total = sum(counts.values())
    print(f"[info] ゼロカーボン都道府県別: 集計件数={total} 内訳={counts}", file=sys.stderr)
    lo, hi = ZERO_CARBON_PLAUSIBLE_RANGE
    if not (lo <= total <= hi):
        print(f"[skip] ゼロカーボン都道府県別: 集計件数 {total} が妥当な範囲外", file=sys.stderr)
        return None
    return counts


def fetch_zero_carbon_pdf_text(pdf_url: str):
    """一覧図PDFを取得し、抽出できたテキストを返す(失敗時はNone)"""
    if pypdf is None:
        print("[skip] ゼロカーボン自治体数: pypdf未インストールのためPDF解析不可", file=sys.stderr)
        return None
    try:
        pdf_raw = fetch(pdf_url)
    except (urllib.error.URLError, TimeoutError, ValueError) as exc:
        print(f"[skip] ゼロカーボン自治体数: PDF取得失敗 ({exc})", file=sys.stderr)
        return None
    try:
        reader = pypdf.PdfReader(io.BytesIO(pdf_raw))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception as exc:  # PDF解析エラーは多岐にわたるため広めに捕捉する
        print(f"[skip] ゼロカーボン自治体数: PDF解析失敗 ({exc})", file=sys.stderr)
        return None


def fetch_zero_carbon_total():
    """環境省サイトから「ゼロカーボンシティ」表明自治体の総数を取得する(失敗時はNone)"""
    try:
        raw = fetch(ZERO_CARBON_URL)
    except (urllib.error.URLError, TimeoutError, ValueError) as exc:
        print(f"[skip] ゼロカーボン自治体数: 取得失敗 ({exc})", file=sys.stderr)
        return None

    html_raw_text = raw.decode("utf-8", errors="ignore")
    total = extract_zero_carbon_total_from_text(strip_html(html_raw_text))
    if total is not None:
        return total

    pdf_url = find_zero_carbon_pdf_url(html_raw_text)
    if not pdf_url:
        print("[skip] ゼロカーボン自治体数: 一覧図PDFへのリンクが見つからず", file=sys.stderr)
        return None

    pdf_text = fetch_zero_carbon_pdf_text(pdf_url)
    if pdf_text:
        total = extract_zero_carbon_total_from_text(pdf_text)
        if total is not None:
            return total
        print(
            f"[debug] ゼロカーボン自治体数: PDFテキスト長={len(pdf_text)} 先頭300文字={pdf_text[:300]!r}",
            file=sys.stderr,
        )

    print("[skip] ゼロカーボン自治体数: 数値を抽出できず", file=sys.stderr)
    return None


def fetch_jepx_spot_average(now: datetime):
    """JEPXスポット市場CSVから直近日のシステムプライス平均(円/kWh)を取得する(失敗時はNone)"""
    for year in (now.year, now.year - 1):
        try:
            body = urllib.parse.urlencode({
                "dir": "spot_summary", "file": f"spot_summary_{year}.csv",
            }).encode("utf-8")
            req = urllib.request.Request(JEPX_DOWNLOAD_URL, data=body, headers={
                "User-Agent": USER_AGENT,
                "Referer": JEPX_SPOT_PAGE_URL,
                "Content-Type": "application/x-www-form-urlencoded",
            })
            with urllib.request.urlopen(req, timeout=FETCH_TIMEOUT) as res:
                raw = res.read()
                content_type = res.headers.get("Content-Type", "")
        except (urllib.error.URLError, TimeoutError, ValueError) as exc:
            print(f"[skip] JEPXスポット価格({year}): 取得失敗 ({exc})", file=sys.stderr)
            continue

        try:
            text = raw.decode("shift_jis", errors="ignore")
            rows = list(csv.DictReader(io.StringIO(text)))
        except csv.Error as exc:
            print(f"[skip] JEPXスポット価格({year}): CSV解析失敗 ({exc})", file=sys.stderr)
            continue
        if not rows:
            print(
                f"[debug] JEPXスポット価格({year}): 行なし content-type={content_type!r} "
                f"先頭300文字={text[:300]!r}",
                file=sys.stderr,
            )
            continue

        date_col = next((k for k in rows[0] if k and ("受渡日" in k or "年月日" in k)), None)
        price_col = next((k for k in rows[0] if k and "システムプライス" in k), None)
        if not date_col or not price_col:
            print(
                f"[debug] JEPXスポット価格({year}): 列一覧={list(rows[0].keys())} "
                f"content-type={content_type!r} 先頭300文字={text[:300]!r}",
                file=sys.stderr,
            )
            print(f"[skip] JEPXスポット価格({year}): 想定する列が見つからず", file=sys.stderr)
            continue

        latest_date = max((r[date_col] for r in rows if r.get(date_col)), default=None)
        if not latest_date:
            continue

        prices = []
        for r in rows:
            if r.get(date_col) != latest_date:
                continue
            try:
                prices.append(float(r[price_col]))
            except (TypeError, ValueError):
                continue
        if prices:
            return round(sum(prices) / len(prices), 1)

    return None


def zero_carbon_collected_recently(now: datetime):
    """ゼロカーボンシティ宣言自治体数(総数・都道府県別)の最終収集日からの経過日数が
    ZERO_CARBON_COLLECT_INTERVAL_DAYS未満ならTrueを返す(変動が少ないため週1回に間引く)"""
    dashboard = load_json("dashboard.json", None) or {}
    as_of = dashboard.get("charts", {}).get("zeroCarbonByPrefecture", {}).get("asOf", "")
    m = ZERO_CARBON_ASOF_RE.search(as_of)
    if not m:
        return False
    y, mo, d = map(int, m.groups())
    try:
        prev = datetime(y, mo, d, tzinfo=now.tzinfo)
    except ValueError:
        return False
    return (now - prev).days < ZERO_CARBON_COLLECT_INTERVAL_DAYS


def update_zero_carbon_kpi(total: int, now: datetime):
    """kpis.json / dashboard.json の「ゼロカーボン宣言」KPIを実数値で更新する"""
    formatted = f"{total:,}"
    source = "環境省（ゼロカーボンシティ表明状況）"
    as_of = f"{now.year}年{now.month}月{now.day}日時点"

    kpis = load_json("kpis.json", [])
    for k in kpis:
        if k.get("label") == "ゼロカーボン宣言":
            prev = int(str(k.get("value", "0")).replace(",", "") or 0)
            diff = total - prev
            k["value"] = formatted
            k["delta"] = f"{diff:+d}"
            k["dir"] = "down" if diff < 0 else "up"
            k["period"] = "前回更新比"
            k["source"] = source
            k["sourceUrl"] = ZERO_CARBON_URL
            k["asOf"] = as_of
    save_json("kpis.json", kpis)

    dashboard = load_json("dashboard.json", None)
    if dashboard is None:
        return
    for k in dashboard.get("kpis", []):
        if k.get("label") == "ゼロカーボン宣言自治体":
            prev = int(str(k.get("value", "0")).replace(",", "") or 0)
            diff = total - prev
            k["value"] = formatted
            k["delta"] = f"{diff:+d}"
            k["dir"] = "down" if diff < 0 else "up"
            k["period"] = "前回更新比"
            k["source"] = source
            k["sourceUrl"] = ZERO_CARBON_URL
            k["asOf"] = as_of
    save_json("dashboard.json", dashboard)


def update_zero_carbon_by_prefecture(counts: dict, now: datetime):
    """dashboard.json の都道府県別ゼロカーボン宣言状況(zeroCarbonByPrefecture)を実数値で更新する"""
    dashboard = load_json("dashboard.json", None)
    if dashboard is None:
        return

    dashboard["zeroCarbonByPrefecture"] = [
        {
            "prefecture": pref,
            "declared": counts.get(pref, 0),
            "total": ZERO_CARBON_MUNICIPALITY_TOTALS[pref],
        }
        for pref in ZERO_CARBON_PREFECTURES
    ]
    charts = dashboard.setdefault("charts", {})
    charts["zeroCarbonByPrefecture"] = {
        "source": "環境省（ゼロカーボンシティ取組一覧）",
        "sourceUrl": ZERO_CARBON_URL,
        "asOf": f"{now.year}年{now.month}月{now.day}日時点",
    }
    save_json("dashboard.json", dashboard)


def update_jepx_price(price: float):
    """rail-data.json の卸電力価格(スポット平均)を実データで更新する"""
    rows = load_json("rail-data.json", [])
    for r in rows:
        if r.get("label") == "卸電力価格":
            r["value"] = f"{price:.1f}"
            r["unit"] = "円/kWh"
    save_json("rail-data.json", rows)


RENEWABLE_XLSX_URL = "https://www.enecho.meti.go.jp/statistics/electric_power/ep002/xls/2025/2-2-2025.xlsx"
RENEWABLE_PAGE_URL = "https://www.enecho.meti.go.jp/statistics/electric_power/ep002/results.html"
# 都道府県別発電実績は月次更新のため、日々の変動が少なく収集を週1回に間引く
RENEWABLE_COLLECT_INTERVAL_DAYS = 7
RENEWABLE_SHEET_YM_RE = re.compile(r"^(\d{4})\.([0-9０-９]+)$")
RENEWABLE_ASOF_RE = re.compile(r"(\d+)年(\d+)月(\d+)日公表時点")


def find_latest_renewable_sheet(sheet_names):
    """月別シート名(例:"2025.4"。全角数字が混じることがある)から最新月のシートを選ぶ"""
    candidates = []
    for name in sheet_names:
        normalized = name.translate(ZENKAKU_DIGITS)
        m = RENEWABLE_SHEET_YM_RE.match(normalized)
        if not m:
            continue
        candidates.append(((int(m.group(1)), int(m.group(2))), name))
    if not candidates:
        return None
    candidates.sort()
    return candidates[-1][1]


def fetch_renewable_ratio_by_prefecture():
    """資源エネルギー庁「都道府県別発電実績」Excelから、都道府県ごとの
    再エネ導入比率((水力+風力+太陽光+地熱)÷合計)を算出する(失敗時はNone)。
    列構成: B=水力 C=火力 D=原子力 E=風力 F=太陽光 G=地熱 H=バイオマス I=廃棄物
    J=蓄電池 K=新エネ計 L=その他 M=合計(いずれも1,000kWh単位)。
    K列(新エネ計)は実際には風力+太陽光+地熱+蓄電池のみで構成されており、
    H列(バイオマス)・I列(廃棄物)はC列(火力発電所)に計上済みの数値の再掲のため、
    重複計上を避けるためbiomass/wasteは合算しない"""
    if openpyxl is None:
        print("[skip] 再エネ導入比率: openpyxl未インストールのためExcel解析不可", file=sys.stderr)
        return None
    try:
        raw = fetch(RENEWABLE_XLSX_URL)
    except (urllib.error.URLError, TimeoutError, ValueError) as exc:
        print(f"[skip] 再エネ導入比率: Excel取得失敗 ({exc})", file=sys.stderr)
        return None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as exc:
        print(f"[skip] 再エネ導入比率: Excel解析失敗 ({exc})", file=sys.stderr)
        return None

    sheet_name = find_latest_renewable_sheet(wb.sheetnames)
    if not sheet_name:
        print(f"[skip] 再エネ導入比率: 月別シートが見つからず (シート一覧={wb.sheetnames})", file=sys.stderr)
        return None

    ws = wb[sheet_name]
    title_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    as_of_text = ""
    for cell in title_row:
        if isinstance(cell, str) and "公表時点" in cell:
            m = RENEWABLE_ASOF_RE.search(cell.translate(ZENKAKU_DIGITS))
            if m:
                y, mo, d = m.groups()
                as_of_text = f"{y}年{mo}月{d}日公表時点"
            break

    results = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        pref = row[0] if row else None
        if pref not in ZERO_CARBON_PREFECTURES:
            continue
        if len(row) < 13:
            continue
        hydro = row[1] or 0
        wind = row[4] or 0
        solar = row[5] or 0
        geo = row[6] or 0
        total = row[12] or 0
        if not total:
            continue
        renewable = hydro + wind + solar + geo
        ratio = round(renewable / total * 100, 1)
        if not (0 <= ratio <= 100):
            print(f"[skip] 再エネ導入比率: {pref}の算出値{ratio}が妥当な範囲外", file=sys.stderr)
            continue
        # total(1,000kWh単位)をkWhに換算して保存。表示側で単位(億kWh等)に丸める
        results.append({"prefecture": pref, "ratio": ratio, "totalKwh": round(total * 1000)})

    print(f"[info] 再エネ導入比率: シート={sheet_name} 集計件数={len(results)}", file=sys.stderr)
    if len(results) < 40:
        print(f"[skip] 再エネ導入比率: 集計件数{len(results)}が少なすぎるため採用しない", file=sys.stderr)
        return None

    return {
        "data": results,
        "source": "資源エネルギー庁「電力調査統計」都道府県別発電実績",
        "sourceUrl": RENEWABLE_PAGE_URL,
        "asOf": as_of_text or f"{sheet_name}分",
    }


def renewable_collected_recently(now: datetime):
    """再エネ導入比率の最終収集日からの経過日数がRENEWABLE_COLLECT_INTERVAL_DAYS未満なら
    Trueを返す(月次更新データのため収集を週1回に間引く)"""
    data = load_json("renewable-by-prefecture.json", None) or {}
    collected_at = data.get("collectedAt", "")
    try:
        prev = datetime.fromisoformat(collected_at)
    except ValueError:
        return False
    return (now - prev).days < RENEWABLE_COLLECT_INTERVAL_DAYS


def collect_subsidies(now: datetime) -> list:
    """環境省・NEDO・経産省の公募・補助金一覧ページからリンクを抽出してsubsidies.json用リストを返す。
    パーサーが想定するHTML構造と実際の構造が異なる場合に備え、診断ログを出力する。"""
    results = []
    seen_urls: set = set()

    for cfg in SUBSIDY_SOURCES_CFG:
        try:
            raw = fetch(cfg["url"])
        except (urllib.error.URLError, TimeoutError, ValueError) as exc:
            print(f"[skip] 補助金収集 {cfg['name']}: 取得失敗 ({exc})", file=sys.stderr)
            continue

        html_text = raw.decode("utf-8", errors="ignore")
        if cfg.get("table_scope"):
            # 共通ナビを含まない本文の<table>部分だけを対象にする
            table_match = re.search(r"<table\b.*?</table>", html_text, re.S | re.I)
            search_text = table_match.group(0) if table_match else ""
        else:
            search_text = html_text

        found: list = []
        for href, raw_title in _SUBSIDY_A_RE.findall(search_text):
            title = strip_html(raw_title).strip()
            if not title or _SUBSIDY_NOISE_RE.match(title):
                continue
            if "href_filter" in cfg and cfg["href_filter"] not in href:
                continue
            # トップや自己参照リンクを除外
            norm = href.rstrip("/")
            if norm in ("", cfg["url"].rstrip("/"), "/"):
                continue
            url = urllib.parse.urljoin(cfg["base_url"], href)
            if url in seen_urls:
                continue
            seen_urls.add(url)
            found.append({
                "title": title,
                "source": cfg["source_label"],
                "url": url,
                "updated": now.strftime("%Y-%m-%d"),
            })

        print(f"[info] 補助金収集 {cfg['name']}: {len(found)}件取得", file=sys.stderr)
        if found:
            print(f"[diag] 補助金サンプル ({cfg['name']}): {found[:2]}", file=sys.stderr)
        else:
            # 0件の場合はページの先頭数百文字を記録して構造調査に役立てる
            print(
                f"[diag] 補助金0件 ({cfg['name']}): ページ先頭400文字={html_text[:400]!r}",
                file=sys.stderr,
            )
        results.extend(found)

    return results


GLOSSARY_PENDING_MAX = 100


def collect_glossary_pending(ranks: list, now: datetime) -> list:
    """話題のキーワードのうち、まだglossary.jsonに登録されていない語を集める。
    定義文の自動生成はできないため、後で人手が確認・登録する前提の候補リストとして
    site/data/glossary_pending.json に保持する(サイト上では件数のみ表示)"""
    glossary = load_json("glossary.json", [])
    existing_terms = {g["term"] for g in glossary}
    # glossaryのtermは「CBAM(炭素国境調整措置)」のように略称+日本語併記の場合があるため、
    # 括弧より前の略称部分でも既存判定する(表記ゆれによる重複登録を避ける)
    existing_prefixes = {t.split("(")[0] for t in existing_terms}

    pending = load_json("glossary_pending.json", [])
    pending_terms = {p["term"] for p in pending}

    now_str = now.strftime("%Y-%m-%d")
    for r in ranks:
        term = r["keyword"]
        if term in existing_terms or term in existing_prefixes or term in pending_terms:
            continue
        pending.append({"term": term, "firstSeen": now_str})
        pending_terms.add(term)

    # 既にglossary.jsonへ登録済みになった語はpendingから外す
    pending = [
        p for p in pending
        if p["term"] not in existing_terms and p["term"] not in existing_prefixes
    ]
    return pending[:GLOSSARY_PENDING_MAX]


def load_json(name: str, default):
    path = SITE_DATA_DIR / name
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def save_json(name: str, data):
    path = SITE_DATA_DIR / name
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def main():
    sources = json.loads(SOURCES_FILE.read_text(encoding="utf-8"))
    now = datetime.now(JST)

    existing_articles = load_json("articles.json", {})
    # 過去にこのスクリプトが収集した記事のみ引き継ぎ、デザイン用サンプル記事は初回実行時に置き換える
    carried_over = {
        aid: a for aid, a in existing_articles.items() if a.get("collected")
    }
    reclassify_carried_over_national(carried_over)

    collected = dict(carried_over)
    new_ids = set()
    for source in sources:
        for article in collect_one_source(source, now):
            prev = collected.get(article["id"])
            if prev is None:
                new_ids.add(article["id"])
            elif "image" in prev:
                # 既知記事を再収集しても、取得済みの画像URLは失わないように引き継ぐ
                article["image"] = prev["image"]
            collected[article["id"]] = article
    new_count = len(new_ids)

    # 同じ記事が複数のRSS/検索クエリ経由で別ID(旧リンクベースIDの引き継ぎ分を含む)として
    # 重複登録されている場合があるため、タイトル単位で1件に統合する
    collected = dedupe_by_title(collected)

    # 新しい順にソートし、上限件数で切る
    ordered_ids = sorted(
        collected, key=lambda k: collected[k].get("_sort_key", ""), reverse=True
    )[:MAX_ARTICLES]
    articles = {aid: collected[aid] for aid in ordered_ids}
    for a in articles.values():
        a.pop("_sort_key", None)

    # 画像は記事ごとに一度だけ取得する(成功・失敗とも結果を保持し、以後は再取得しない)
    for article in articles.values():
        if "image" not in article:
            article["image"] = fetch_article_image(article["source_url"])

    feed = [
        {
            "id": aid,
            "title": articles[aid]["title"],
            "category": articles[aid]["category"],
            "accent": i % 3 == 1,
            "source": articles[aid]["source"],
            "time": articles[aid]["time"].split(" ")[-1],
            "source_url": articles[aid]["source_url"],
            "image": articles[aid].get("image"),
        }
        for i, aid in enumerate(ordered_ids[:MAX_FEED_ITEMS])
    ]

    category_counts = {}
    for a in articles.values():
        category_counts[a["category"]] = category_counts.get(a["category"], 0) + 1

    categories = load_json("categories.json", [])
    for c in categories:
        c["count"] = category_counts.get(c["label"], 0)

    previous_ranks = load_json("ranks.json", [])
    topics = build_topics(articles, ordered_ids, categories, new_ids)
    ranks = build_ranks(articles, previous_ranks)
    previous_events = load_json("events.json", [])
    events = build_events(articles, previous_events, now)

    save_json("articles.json", articles)
    save_json("feed.json", feed)
    save_json("categories.json", categories)
    if topics is not None:
        save_json("topics.json", topics)
    if ranks is not None:
        save_json("ranks.json", ranks)
        save_json("glossary_pending.json", collect_glossary_pending(ranks, now))
    save_json("events.json", events)
    save_json("meta.json", {"updated_at": now.isoformat()})

    if zero_carbon_collected_recently(now):
        print(f"[skip] ゼロカーボンシティ宣言自治体: 前回収集から{ZERO_CARBON_COLLECT_INTERVAL_DAYS}日未満のため収集をスキップ", file=sys.stderr)
    else:
        zero_carbon_total = fetch_zero_carbon_total()
        if zero_carbon_total is not None:
            update_zero_carbon_kpi(zero_carbon_total, now)

        zero_carbon_by_prefecture = fetch_zero_carbon_by_prefecture()
        if zero_carbon_by_prefecture is not None:
            update_zero_carbon_by_prefecture(zero_carbon_by_prefecture, now)

    jepx_price = fetch_jepx_spot_average(now)
    if jepx_price is not None:
        update_jepx_price(jepx_price)

    subsidies = collect_subsidies(now)
    save_json("subsidies.json", subsidies)

    if renewable_collected_recently(now):
        print(f"[skip] 再エネ導入比率: 前回収集から{RENEWABLE_COLLECT_INTERVAL_DAYS}日未満のため収集をスキップ", file=sys.stderr)
    else:
        renewable_result = fetch_renewable_ratio_by_prefecture()
        if renewable_result is not None:
            renewable_result["collectedAt"] = now.isoformat()
            save_json("renewable-by-prefecture.json", renewable_result)

    print(f"収集完了: 新規 {new_count} 件 / 合計 {len(articles)} 件 ({now.isoformat()})")


if __name__ == "__main__":
    SITE_DATA_DIR.mkdir(parents=True, exist_ok=True)
    main()
